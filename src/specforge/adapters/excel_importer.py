from __future__ import annotations

from pathlib import Path

from pydantic import ValidationError

from specforge.adapters.csv_mapper import build_spec_tree
from specforge.adapters.csv_schema import (
    CSVFieldRow,
    CSVImportError,
    _parse_row,
    _validate_headers,
)
from specforge.models.spec import SpecFile

_HEADERS = (
    "field_name", "type", "required", "nullable", "description",
    "item_type", "format", "enum", "default",
    "min_length", "max_length", "pattern",
    "minimum", "maximum",
    "min_items", "max_items", "unique_items",
)

_MAX_EXCEL_BYTES = 25 * 1024 * 1024  # 25 MB
_MAX_EXCEL_ROWS = 100_000


def import_excel(path: Path, sheet: str | None = None) -> SpecFile:
    try:
        import openpyxl
    except ImportError as exc:
        raise CSVImportError("openpyxl is required for Excel import: pip install openpyxl") from exc

    size = path.stat().st_size
    if size > _MAX_EXCEL_BYTES:
        raise CSVImportError(
            f"Excel file is {size} bytes; maximum allowed is {_MAX_EXCEL_BYTES} bytes"
        )

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise CSVImportError(f"Could not open Excel file: {exc}") from exc

    try:
        rows = _extract_rows(wb, sheet)
    finally:
        wb.close()

    if not rows:
        raise CSVImportError(
            "Excel file has a header but no field rows — at least one non-blank data row is required"
        )

    try:
        spec_data = build_spec_tree(rows)
        return SpecFile.model_validate(spec_data)
    except ValidationError as exc:
        raise CSVImportError(f"Generated spec is invalid:\n{exc}") from exc


def _extract_rows(wb, sheet: str | None) -> list[CSVFieldRow]:
    if sheet is None:
        ws = wb.active
    else:
        if sheet not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise CSVImportError(
                f"Sheet '{sheet}' not found in workbook (available: {available})"
            )
        ws = wb[sheet]

    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise CSVImportError("Excel file is empty")

    indexed_headers: list[tuple[int, str]] = []
    for index, raw_header in enumerate(header_row):
        if raw_header is None:
            continue
        name = str(raw_header).strip()
        if name:
            indexed_headers.append((index, name))

    if not indexed_headers:
        raise CSVImportError("Excel file is empty")

    headers = [name for _, name in indexed_headers]
    _validate_headers(headers)

    rows: list[CSVFieldRow] = []
    for row_num, raw_row in enumerate(rows_iter, start=2):
        if row_num - 1 > _MAX_EXCEL_ROWS:
            raise CSVImportError(
                f"Excel file exceeds row limit of {_MAX_EXCEL_ROWS}"
            )

        normalized: dict[str, str | None] = {}
        for col_index, header in indexed_headers:
            cell_val = raw_row[col_index] if col_index < len(raw_row) else None
            if cell_val is None:
                normalized[header] = None
            else:
                s = str(cell_val).strip()
                normalized[header] = s if s else None

        parsed = _parse_row(row_num, normalized)
        if parsed is not None:
            rows.append(parsed)

    return rows
