from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import openpyxl
import pytest

from specforge.adapters import CSVImportError, import_excel
from specforge.models.spec import SpecFile

HEADERS = (
    "field_name", "type", "required", "nullable", "description",
    "item_type", "format", "enum", "default",
    "min_length", "max_length", "pattern",
    "minimum", "maximum",
    "min_items", "max_items", "unique_items",
)


def _write_xlsx(name: str, rows: list[dict]) -> Path:
    path = Path(f"excel-importer-{uuid4().hex}-{name}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(HEADERS))
    for row in rows:
        ws.append([row.get(h, None) for h in HEADERS])
    wb.save(path)
    return path


def test_minimal_excel_maps_to_specfile() -> None:
    path = _write_xlsx("minimal", [{"field_name": "name", "type": "string", "required": True, "nullable": False}])
    spec = import_excel(path)

    assert isinstance(spec, SpecFile)
    assert spec.fields["name"].type == "string"
    assert spec.fields["name"].required is True
    assert spec.fields["name"].nullable is False


def test_missing_required_header_fails() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["field_name", "required"])
    ws.append(["name", "true"])
    path = Path(f"excel-importer-{uuid4().hex}-missing-header.xlsx")
    wb.save(path)

    with pytest.raises(CSVImportError, match="Missing required CSV header"):
        import_excel(path)


def test_unknown_extra_column_fails() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["field_name", "type", "wat"])
    ws.append(["name", "string", "x"])
    path = Path(f"excel-importer-{uuid4().hex}-extra-header.xlsx")
    wb.save(path)

    with pytest.raises(CSVImportError, match="Unknown CSV header"):
        import_excel(path)


def test_empty_xlsx_fails() -> None:
    wb = openpyxl.Workbook()
    path = Path(f"excel-importer-{uuid4().hex}-empty.xlsx")
    wb.save(path)

    with pytest.raises(CSVImportError, match="Excel file is empty"):
        import_excel(path)


def test_header_only_xlsx_fails() -> None:
    path = _write_xlsx("header-only", [])

    with pytest.raises(CSVImportError, match="at least one non-blank data row"):
        import_excel(path)


def test_blank_rows_are_ignored() -> None:
    path = _write_xlsx("blank-rows", [
        {},
        {"field_name": "name", "type": "string"},
    ])
    spec = import_excel(path)

    assert set(spec.fields) == {"name"}


def test_numeric_cell_values_are_coerced() -> None:
    path = _write_xlsx("numeric", [
        {"field_name": "tags", "type": "array", "item_type": "string", "min_items": 1, "max_items": 5},
    ])
    spec = import_excel(path)

    assert spec.fields["tags"].minItems == 1
    assert spec.fields["tags"].maxItems == 5


def test_boolean_cell_values_are_coerced() -> None:
    path = _write_xlsx("bool-cells", [
        {"field_name": "name", "type": "string", "required": True, "nullable": False},
    ])
    spec = import_excel(path)

    assert spec.fields["name"].required is True
    assert spec.fields["name"].nullable is False


def test_enum_pipe_split_works() -> None:
    path = _write_xlsx("enum", [
        {"field_name": "status", "type": "string", "required": True, "nullable": False, "enum": "A|B|C"},
    ])
    spec = import_excel(path)

    assert spec.fields["status"].enum == ["A", "B", "C"]


def test_dot_notation_creates_nested_object() -> None:
    path = _write_xlsx("nested", [
        {"field_name": "address.street", "type": "string", "required": True, "nullable": False},
        {"field_name": "address.city", "type": "string"},
    ])
    spec = import_excel(path)

    assert spec.fields["address"].type == "object"
    assert "street" in spec.fields["address"].fields
    assert "city" in spec.fields["address"].fields


def test_object_array_items_fields() -> None:
    path = _write_xlsx("object-array", [
        {"field_name": "items", "type": "array", "required": True, "nullable": False, "item_type": "object"},
        {"field_name": "items.item.code", "type": "string", "required": True, "nullable": False},
    ])
    spec = import_excel(path)

    assert spec.fields["items"].items is not None
    assert spec.fields["items"].items.type == "object"
    assert "code" in spec.fields["items"].items.fields


def test_generated_structure_validates_against_specfile_model() -> None:
    path = _write_xlsx("valid-structure", [
        {"field_name": "id", "type": "integer", "required": True, "nullable": False},
        {"field_name": "name", "type": "string"},
        {"field_name": "tags", "type": "array", "item_type": "string"},
    ])
    spec = import_excel(path)

    assert isinstance(spec, SpecFile)
